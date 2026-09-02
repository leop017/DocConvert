from __future__ import annotations

import re
from typing import Any

INVALID_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}


def get_excel_sheet_names(filepath: str, ext: str) -> list[str]:
    """Return the list of sheet names in an Excel workbook.

    ``ext`` is the lowercased suffix (``.xls`` or ``.xlsx``); ``.xls`` is
    read via xlrd and ``.xlsx`` via openpyxl.
    """
    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(filepath)
        try:
            return wb.sheet_names()
        finally:
            wb.release_resources()
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and str(value) == 'nan':
        return ""
    try:
        import pandas as pd
        if pd.isna(value):
            return ""
    except (ImportError, TypeError, ValueError):
        pass
    s = str(value)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    return s


def clean_filename(name: str) -> str:
    name = re.sub(r'[\x00-\x1f]', '', str(name))
    name = re.sub(r'[\/\\:*?"<>|]', "_", name)
    if not name.strip():
        name = "untitled"
    stem = name.split('.')[0].upper()
    if stem in INVALID_NAMES:
        name = f"_{name}"
    if len(name) > 180:
        name = name.encode('utf-8')[:180].decode('utf-8', errors='replace')
    return name


def unique_cleaned_suffixes(names: list[str]) -> list[str]:
    """Return a clean_filename-escaped, batch-unique suffix per name.

    Inputs that differ but clean to the same value (e.g. sheet names
    ``Q"1`` and ``Q<1`` both become ``Q_1``) would otherwise produce
    identical output paths and the later item would silently overwrite
    the earlier one. Subsequent collisions get a numeric suffix:
    ``name``, ``name_2``, ``name_3``...
    """
    used: set[str] = set()
    result: list[str] = []
    for name in names:
        base = clean_filename(name)
        candidate = base
        n = 2
        while candidate in used:
            candidate = f'{base}_{n}'
            n += 1
        used.add(candidate)
        result.append(candidate)
    return result


def decode_text(raw: bytes) -> str:
    """Decode byte output from external tools (textract/antiword).

    Their output encoding depends on the document and the system locale
    (UTF-8 on most systems, GBK on zh-CN Windows, Latin-1 on Western).
    Try UTF-8 first, then GB18030 (a superset of GBK/GB2312), then fall
    back to Latin-1, which can never fail.
    """
    for enc in ('utf-8', 'gb18030', 'latin-1'):
        try:
            s = raw.decode(enc)
            if '\ufffd' not in s:
                return s
        except (UnicodeDecodeError, ValueError):
            continue
    return raw.decode('latin-1', errors='replace')


def escape_md_cell(text: str) -> str:
    """Escape Markdown-significant characters inside a table cell.

    Markdown table cells containing a literal ``|`` break the row
    layout because the pipe is also the column separator. Cells
    containing a newline are joined by ``markdownify`` into a
    single space, losing the visible line break. This helper returns
    a Markdown-safe form of the cell text so the rendered table
    keeps its column count and the line break is preserved as
    ``<br>`` (which survives markdownify pass-through).

    Apply this only to text that will be embedded inside a ``<th>``
    / ``<td>``. Do not apply it to the surrounding HTML structure.
    """
    if not text:
        return text
    return text.replace("|", "\\|").replace("\n", "<br>")

def html_to_md(html_content: str) -> str:
    import markdownify
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')
    for tag in soup(['script', 'style']):
        tag.extract()
    return markdownify.markdownify(str(soup), heading_style=markdownify.ATX).strip()
