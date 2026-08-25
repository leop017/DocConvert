from __future__ import annotations

import html as html_mod
import json
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from docconvert.config import AppConfig
from docconvert.converters.base import BaseConverter
from docconvert.models import MergeInfo
from docconvert.utils import (
    clean_filename,
    escape_md_cell,
    html_to_md,
    safe_str,
    unique_cleaned_suffixes,
)

class _XlsMergeRange:
    """Mimics openpyxl's merged cell range for use with _build_merged_map."""
    def __init__(self, min_col: int, min_row: int, max_col: int, max_row: int):
        self.bounds = (min_col, min_row, max_col, max_row)


class ExcelConverter(BaseConverter):

    def __init__(self, config: Optional[AppConfig] = None):
        super().__init__(config)

    def convert(
        self,
        input_path: str,
        output_fmt: str,
        parent: Path,
        **kwargs,
    ) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
        results: list[tuple[str, str]] = []
        errors: list[tuple[str, str]] = []

        ext = Path(input_path).suffix.lower()
        engine = 'xlrd' if ext == '.xls' else 'openpyxl'

        if ext == '.xls':
            try:
                import xlrd
            except ImportError:
                raise RuntimeError('需安装 xlrd 库以处理 .xls 文件')

        sheets_param = kwargs.get('sheets')
        enhanced_md = kwargs.get('enhanced_md', False)
        stem_override = kwargs.get('stem_override')

        if sheets_param is not None:
            all_sheet_names = self._get_all_sheet_names(input_path, ext)
            existing = [sn for sn in sheets_param if sn in all_sheet_names]
            missing = [sn for sn in sheets_param if sn not in all_sheet_names]
            for sn in missing:
                errors.append((sn, '工作表不存在'))
            if not existing:
                return results, errors
            all_sheets = pd.read_excel(
                input_path, sheet_name=existing, engine=engine
            )
            if not isinstance(all_sheets, dict):
                all_sheets = {existing[0]: all_sheets}
            # Only load merged-cell metadata for sheets we will actually
            # convert — avoids the (small but real) overhead of opening
            # the workbook a second time for unrequested sheets.
            sheet_names = existing
            sheets_to_convert = existing
        else:
            all_sheets = pd.read_excel(input_path, sheet_name=None, engine=engine)
            sheet_names = list(all_sheets.keys())
            sheets_to_convert = sheet_names

        merged_cache = self._load_merged_cache(input_path, ext, sheet_names)

        # Sheet names that clean to the same value (e.g. ``Q"1`` and
        # ``Q<1`` → ``Q_1``) must not share an output path, or the later
        # sheet would silently overwrite the earlier one. Resolve unique
        # suffixes up front so the emitted filenames are unambiguous.
        sn_overrides = dict(
            zip(sheets_to_convert, unique_cleaned_suffixes(sheets_to_convert))
        )

        for sn in sheets_to_convert:
            if self.cancel_check and self.cancel_check():
                self.logger.info("转换已取消")
                break
            try:
                df = all_sheets.get(sn)
                if df is None:
                    errors.append((sn, '工作表不存在'))
                    continue
                mr = merged_cache.get(sn) if merged_cache else None
                name, path = self._convert_sheet(
                    input_path, df, sn, output_fmt, parent, mr,
                    enhanced_md=enhanced_md, stem_override=stem_override,
                    sn_override=sn_overrides.get(sn),
                )
                results.append((name, path))
            except Exception as e:
                self.logger.error("工作表转换失败 [%s]: %s", sn, str(e))
                errors.append((sn, str(e)))

        return results, errors

    def _load_merged_cache(
        self, input_path: str, ext: str, sheet_names: list[str]
    ) -> Optional[dict[str, list]]:
        if ext == '.xls':
            return self._load_merged_cache_xls(input_path, sheet_names)
        if ext != '.xlsx':
            return None
        try:
            wb = load_workbook(input_path, data_only=True)
        except Exception as e:
            self.logger.warning("打开工作簿失败： %s", e)
            return None
        try:
            cache: dict[str, list] = {}
            for sn in sheet_names:
                try:
                    cache[sn] = list(wb[sn].merged_cells.ranges)
                except (KeyError, AttributeError) as e:
                    self.logger.debug("工作表 '%s' 合并单元格读取失败： %s", sn, e)
                    cache[sn] = []
            return cache
        except Exception as e:
            self.logger.warning("合并单元格读取失败： %s", e)
            return None
        finally:
            wb.close()

    @staticmethod
    def _get_all_sheet_names(input_path: str, ext: str) -> list[str]:
        from docconvert.utils import get_excel_sheet_names
        return get_excel_sheet_names(input_path, ext)

    @staticmethod
    def _load_merged_cache_xls(
        input_path: str, sheet_names: list[str]
    ) -> dict[str, list]:
        import xlrd
        wb = xlrd.open_workbook(input_path, formatting_info=True)
        try:
            cache: dict[str, list] = {}
            for sn in sheet_names:
                try:
                    ws = wb.sheet_by_name(sn)
                except (KeyError, xlrd.XLRDError):
                    cache[sn] = []
                    continue
                merged = []
                for rlo, rhi, clo, chi in ws.merged_cells:
                    merged.append(_XlsMergeRange(clo + 1, rlo + 1, chi, rhi))
                cache[sn] = merged
            return cache
        finally:
            wb.release_resources()

    def _build_merged_map(self, merged_ranges: list) -> dict:
        merged_map = {}
        for merged in merged_ranges:
            min_col, min_row, max_col, max_row = merged.bounds
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    key = (row, col)
                    if key not in merged_map:
                        merged_map[key] = MergeInfo(
                            rowspan=max_row - min_row + 1,
                            colspan=max_col - min_col + 1,
                            is_master=(row == min_row and col == min_col),
                            is_merged=True,
                            min_row=min_row,
                            min_col=min_col,
                            max_row=max_row,
                            max_col=max_col,
                        )
                    else:
                        merged_map[key].is_master = False
        return merged_map

    def _df_to_rows(self, df: pd.DataFrame) -> list[list[str]]:
        rows_data = []
        rows_data.append([safe_str(c) for c in df.columns.tolist()])
        for _, row in df.iterrows():
            rows_data.append([safe_str(v) for v in row.tolist()])
        return rows_data

    @staticmethod
    def _filter_merges_for_dropped_rows(
        merged_ranges: list, dropped_df_indices: set
    ) -> list:
        """Filter merge ranges against the rows that ``dropna`` removed.

        ``merged_ranges`` use 1-based workbook row numbers; ``pd.read_excel``
        consumes the first row as the header, so workbook row N (N >= 2)
        maps to DataFrame index (N - 2). The header row (workbook row 1)
        has no DataFrame index (its formula ``row - 2`` yields ``-1``) and
        is never dropped.

        A merge is kept only when none of its rows landed in
        ``dropped_df_indices``. If any leg was dropped, the surviving rows
        below the gap shift upward and no longer align with the merge's
        original row coordinates, so keeping the merge would mask a real
        cell (silent data loss) or visually jump over the gap.

        This uniformly covers header-anchored merges (``min_row == 1``):
        their master lives in the consumed-header region, but their
        vertical legs can still collide with a dropped interior row — e.g.
        ``A1:A3`` with an empty row 3 removed by ``dropna`` would otherwise
        make the master's rowspan swallow the independent ``A4`` cell that
        shifted up into its place.
        """
        filtered: list = []
        for merged in merged_ranges:
            min_col, min_row, max_col, max_row = merged.bounds
            spans_dropped = any(
                (row - 2) in dropped_df_indices
                for row in range(min_row, max_row + 1)
            )
            if not spans_dropped:
                filtered.append(merged)
        return filtered

    @staticmethod
    def _remap_merged_rows(merged_ranges: list, df: "pd.DataFrame") -> list:
        """Translate workbook row numbers to post-``dropna`` rendered rows.

        ``merged_ranges`` arrive in 1-based workbook coordinates. After
        ``dropna`` removes fully-empty rows, the surviving data rows are
        re-indexed sequentially (workbook row R maps to rendered row
        ``2 + position``). Build that mapping and re-base every merge's
        row bounds onto the rendered coordinates so the merge lands on the
        correct cells.

        A merge that survives filtering cannot span a dropped row, so all
        of its rows exist in the mapping; if one does not (defensive),
        drop the merge rather than render it misaligned.
        """
        wb_to_rendered = {1: 1}
        for pos, orig_idx in enumerate(df.index):
            wb_to_rendered[orig_idx + 2] = 2 + pos
        remapped: list = []
        for merged in merged_ranges:
            min_col, min_row, max_col, max_row = merged.bounds
            if min_row not in wb_to_rendered or max_row not in wb_to_rendered:
                continue
            remapped.append(
                _XlsMergeRange(
                    min_col,
                    wb_to_rendered[min_row],
                    max_col,
                    wb_to_rendered[max_row],
                )
            )
        return remapped

    @staticmethod
    def _iter_merge_legs(info):
        """Yield every (row, col) covered by info.

        Used by the merge guard in _convert_sheet to walk every cell
        a merge occupies, including its slave legs.
        """
        for row in range(info.min_row, info.max_row + 1):
            for col in range(info.min_col, info.max_col + 1):
                yield (row, col)

    def _convert_sheet(
        self,
        input_path: str,
        df: pd.DataFrame,
        sheet_name: str,
        output_fmt: str,
        parent: Path,
        merged_ranges: Optional[list] = None,
        enhanced_md: bool = False,
        stem_override: Optional[str] = None,
        sn_override: Optional[str] = None,
    ) -> tuple[str, str]:
        if df.empty:
            raise ValueError('工作表为空')

        all_nan_mask = df.isna().all(axis=1)
        dropped_df_indices = set(df.index[all_nan_mask].tolist())
        if merged_ranges and dropped_df_indices:
            merged_ranges = self._filter_merges_for_dropped_rows(
                merged_ranges, dropped_df_indices
            )

        df = df.dropna(how='all')
        if df.empty:
            raise ValueError('工作表为空')

        rows_data = self._df_to_rows(df)
        max_cols = len(df.columns)
        merged_map = {}
        if merged_ranges:
            # ``merged_ranges`` carry original 1-based workbook row numbers,
            # but ``_df_to_rows`` re-indexes the surviving rows sequentially
            # (header -> rendered row 1, data rows -> 2, 3, ...). Any fully
            # empty row removed by ``dropna`` shifts every later row up, so
            # the workbook coordinates no longer line up with the rendered
            # table. Remap them before building ``merged_map`` or merges
            # after a blank row would render at the wrong row (or swallow
            # the wrong cell's value).
            merged_ranges = self._remap_merged_rows(merged_ranges, df)
            merged_map = self._build_merged_map(merged_ranges)
            mc = max((key[1] for key in merged_map), default=0)
            max_cols = max(max_cols, mc)
            # Bug guard: a merge whose legs extend past the available
            # rows would render a rowspan that no tbody row covers,
            # so the browser silently clamps the cell and the visual
            # shape becomes wrong. ``pd.read_excel`` silently drops
            # trailing fully-empty workbook rows, so the dropna-based
            # filter above cannot detect this case. Drop the merge
            # entirely when any leg is past the rendered rows.
            rows_after_dropna = len(rows_data)
            for key, info in list(merged_map.items()):
                if info.max_row > rows_after_dropna:
                    master_key = (info.min_row, info.min_col)
                    for leg in self._iter_merge_legs(info):
                        merged_map.pop(leg, None)
                    merged_map.pop(master_key, None)


        stem = stem_override or clean_filename(Path(input_path).stem)
        sn_clean = sn_override or clean_filename(sheet_name)
        source_name = Path(input_path).name

        if output_fmt == 'html':
            return self._write_html(rows_data, merged_map, max_cols, sheet_name, stem, sn_clean, source_name, parent)
        elif output_fmt == 'md':
            return self._write_md(df, rows_data, merged_map, max_cols, sheet_name, stem, sn_clean, source_name, parent, enhanced_md)
        elif output_fmt == 'json':
            return self._write_json(rows_data, merged_map, sheet_name, stem, sn_clean, parent)
        raise ValueError(f'不支持的输出格式： {output_fmt}')

    # ── HTML output ──────────────────────────────────────────────

    def _write_html(
        self, rows_data: list, merged_map: dict, max_cols: int,
        sheet_name: str, stem: str, sn_clean: str, source_name: str, parent: Path,
    ) -> tuple[str, str]:
        html_table = self._build_html_table(rows_data, merged_map, max_cols, sheet_name, source_name)
        full_html = _html_document(html_table, sheet_name)
        output_name = f'{stem}_{sn_clean}.html'
        output_path = str(parent / output_name)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(self._export(full_html, 'html'))
        return output_name, output_path

    def _build_html_table(
        self, rows_data: list, merged_map: dict, max_cols: int, sheet_name: str,
        source_name: str,
    ) -> str:
        header_row = rows_data[0] if rows_data else []
        html_rows = []
        header_cells = []
        # Cells in the body covered by a header-anchored vertical merge.
        # A ``<th rowspan>`` cannot reach from <thead> into <tbody>, so such
        # merges are flattened: the header renders a normal cell and the
        # covered body cells become placeholders, keeping the table aligned.
        header_covered: dict = {}

        for col_pos in range(1, max_cols + 1):
            h = header_row[col_pos - 1] if col_pos <= len(header_row) else ''
            key = (1, col_pos)
            cell_info = merged_map.get(key)
            if cell_info and cell_info.is_master:
                if cell_info.max_row > 1:
                    for r in range(1, cell_info.rowspan):
                        for c in range(cell_info.colspan):
                            header_covered[(1 + r, col_pos + c)] = True
                    # A header-anchored merge that also spans columns still
                    # needs its colspan so the thead keeps the same column
                    # count as the tbody; rowspan alone cannot cross the
                    # thead/tbody boundary, so it is flattened (see above).
                    span = (
                        f' colspan="{cell_info.colspan}"'
                        f' data-colspan="{cell_info.colspan}"'
                        if cell_info.colspan > 1 else ''
                    )
                    cell_value = html_mod.escape(str(h)) if h is not None and str(h).strip() != '' else '&nbsp;'
                    header_cells.append(f'<th scope="col" data-row="1" data-col="{col_pos}"{span}>{cell_value}</th>')
                else:
                    attrs = self._cell_attrs(1, col_pos, cell_info)
                    cell_value = html_mod.escape(str(h)) if h is not None and str(h).strip() != '' else '&nbsp;'
                    header_cells.append(f'<th{attrs}>{cell_value}</th>')
            elif cell_info and not cell_info.is_master:
                continue
            else:
                cell_value = html_mod.escape(str(h)) if h is not None and str(h).strip() != '' else '&nbsp;'
                header_cells.append(f'<th scope="col" data-row="1" data-col="{col_pos}">{cell_value}</th>')
        html_rows.append('<tr data-row="1">' + ''.join(header_cells) + '</tr>')

        row_spans: dict = {}
        for row_idx, row_data in enumerate(rows_data[1:], start=2):
            cells = []
            for col in range(1, max_cols + 1):
                if header_covered.get((row_idx, col)):
                    cells.append('<td>&nbsp;</td>')
                    continue
                if row_spans.get((row_idx, col), False):
                    continue
                value = row_data[col - 1] if col <= len(row_data) else ''
                key = (row_idx, col)
                cell_info = merged_map.get(key)
                if cell_info and not cell_info.is_master:
                    continue
                if cell_info and cell_info.is_master:
                    attrs = self._cell_attrs(row_idx, col, cell_info)
                    for r in range(1, cell_info.rowspan):
                        for c in range(cell_info.colspan):
                            row_spans[(row_idx + r, col + c)] = True
                    cell_value = html_mod.escape(str(value)) if value is not None and str(value).strip() != '' else '&nbsp;'
                    cells.append(f'<td{attrs}>{cell_value}</td>')
                else:
                    cell_value = html_mod.escape(str(value)) if value is not None and str(value).strip() != '' else '&nbsp;'
                    cells.append(f'<td>{cell_value}</td>')
            html_rows.append(f'<tr data-row="{row_idx}">' + ''.join(cells) + '</tr>')

        source_name = html_mod.escape(source_name)
        return (
            f'<table border="1" class="excel-data" id="data-table" '
            f'data-sheet="{html_mod.escape(sheet_name)}"'
            f' data-source="{source_name}">\n<thead>\n{html_rows[0]}\n</thead>\n<tbody>\n'
            + '\n'.join(html_rows[1:]) + '\n</tbody>\n</table>'
        )

    @staticmethod
    def _cell_attrs(row_idx: int, col_idx: int, info: MergeInfo) -> str:
        parts = [f'data-row="{row_idx}"', f'data-col="{col_idx}"']
        if info.rowspan > 1:
            parts.append(f'rowspan="{info.rowspan}"')
            parts.append(f'data-rowspan="{info.rowspan}"')
        if info.colspan > 1:
            parts.append(f'colspan="{info.colspan}"')
            parts.append(f'data-colspan="{info.colspan}"')
        return ' ' + ' '.join(parts)

    # ── Markdown output (via markitdown) ─────────────────────────

    def _write_md(
        self, df: pd.DataFrame, rows_data: list, merged_map: dict, max_cols: int,
        sheet_name: str, stem: str, sn_clean: str, source_name: str, parent: Path, enhanced: bool,
    ) -> tuple[str, str]:
        if enhanced:
            md_content = self._generate_md_via_html(rows_data, merged_map, max_cols, sheet_name)
        else:
            md_content = self._generate_md_standard(df, sheet_name)

        md_content = (
            f'<!-- source: {source_name} | sheet: {sheet_name}'
            f' | rows: {len(rows_data) - 1} | cols: {max_cols} -->\n\n{md_content}'
        )
        output_name = f'{stem}_{sn_clean}.md'
        output_path = str(parent / output_name)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(self._export(md_content, 'md'))
        return output_name, output_path

    def _generate_md_standard(self, df: pd.DataFrame, sheet_name: str) -> str:
        html_content = df.to_html(index=False, na_rep='')
        return html_to_md(self._escape_table_cells(html_content))

    @staticmethod
    def _escape_table_cells(html_content: str) -> str:
        """Apply escape_md_cell to every th / td text node.

        df.to_html escapes newlines as a literal backslash-n
        sequence (a 2-char run) so HTML renderers display them as
        text. For Markdown output that is wrong: we want the cell
        to render a real line break (``<br>``). Undo pandas'
        escaping before applying escape_md_cell, so the Markdown
        pipeline sees an actual newline and emits ``<br>``.

        Only the standard MD path runs through here. The enhanced
        MD path feeds rows_data directly into the HTML builder,
        where newlines are never escaped by pandas.
        """
        from bs4 import BeautifulSoup, Tag
        soup = BeautifulSoup(html_content, 'html.parser')
        for cell in soup.find_all(['th', 'td']):
            if not isinstance(cell, Tag):
                continue
            current = cell.get_text()
            unescaped = current.replace(chr(92) + chr(110), chr(10))
            escaped = escape_md_cell(unescaped)
            if escaped != current:
                cell.string = escaped
        return str(soup)

    def _generate_md_via_html(
        self, rows_data: list, merged_map: dict, max_cols: int, sheet_name: str,
    ) -> str:
        row_spans: dict = {}
        html_rows = ['<table>']
        for row_idx, row_data in enumerate(rows_data, start=1):
            cells = []
            for col in range(1, max_cols + 1):
                span_val = row_spans.get((row_idx, col))
                if span_val is not None:
                    cells.append(span_val)
                    continue
                value = row_data[col - 1] if col <= len(row_data) else ''
                key = (row_idx, col)
                cell_info = merged_map.get(key)
                tag = 'th' if row_idx == 1 else 'td'
                if cell_info and not cell_info.is_master:
                    cells.append(f'<{tag}>&nbsp;</{tag}>')
                    continue
                if cell_info and cell_info.is_master:
                    if value is not None and str(value).strip() != '':
                        # escape_md_cell turns a real newline into a literal
                        # ``<br>``; escaping afterwards keeps it as text
                        # (``&lt;br&gt;``) so the intermediate HTML parsed by
                        # ``html_to_md`` does not turn it into a real element
                        # that markdownify then collapses to a space. This
                        # mirrors the standard-MD path in ``_escape_table_cells``.
                        cell_value = html_mod.escape(escape_md_cell(str(value)))
                    else:
                        cell_value = '&nbsp;'
                    cells.append(f'<{tag}>{cell_value}</{tag}>')
                    for r in range(cell_info.rowspan):
                        for c in range(cell_info.colspan):
                            if r == 0 and c == 0:
                                continue
                            row_spans[(row_idx + r, col + c)] = f'<{tag}>{cell_value}</{tag}>'
                else:
                    if value is not None and str(value).strip() != '':
                        cell_value = html_mod.escape(escape_md_cell(str(value)))
                    else:
                        cell_value = '&nbsp;'
                    cells.append(f'<{tag}>{cell_value}</{tag}>')
            html_rows.append('<tr>' + ''.join(cells) + '</tr>')
        html_rows.append('</table>')
        return html_to_md('\n'.join(html_rows))

    # ── JSON output ──────────────────────────────────────────────

    def _write_json(
        self, rows_data: list, merged_map: dict, sheet_name: str,
        stem: str, sn_clean: str, parent: Path,
    ) -> tuple[str, str]:
        data = _generate_json_data(rows_data, merged_map, sheet_name)
        output_name = f'{stem}_{sn_clean}.json'
        output_path = str(parent / output_name)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(self._export(data, 'json'))
        return output_name, output_path


# ── Module-level helpers ─────────────────────────────────────────


def _html_document(table_html: str, sheet_name: str) -> str:
    return (
        f'<!DOCTYPE html>\n<html lang="zh-CN" data-exported-by="DocConvert" '
        f'data-sheet-name="{html_mod.escape(sheet_name)}">\n'
        f'<head>\n    <meta charset="UTF-8">\n    '
        f'<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
        f'    <title>{html_mod.escape(sheet_name)}</title>\n'
        f'    <style>\n'
        f'        body {{ font-family: Arial, sans-serif; margin: 20px; }}\n'
        f'        table {{ border-collapse: collapse; width: 100%; }}\n'
        f'        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; white-space: pre-wrap; }}\n'
        f'        th {{ background-color: #4CAF50; color: white; }}\n'
        f'        tr:nth-child(even) {{ background-color: #f2f2f2; }}\n'
        f'        tbody tr {{ cursor: pointer; }}\n'
        f'        tbody tr:hover {{ background-color: #e0e0e0; }}\n'
        f'    </style>\n</head>\n<body>\n{table_html}\n</body>\n</html>'
    )


def _generate_json_data(rows_data: list, merged_map: dict, sheet_name: str) -> str:
    if not rows_data or len(rows_data) < 2:
        return json.dumps(
            {'metadata': {'sheet': sheet_name, 'total_rows': 0, 'total_columns': 0, 'headers': [], 'merged_cells': []}, 'data': []},
            ensure_ascii=False, indent=2,
        )

    headers = list(rows_data[0])
    for c in range(len(headers)):
        if not headers[c]:
            for r in range(1, len(rows_data)):
                if c < len(rows_data[r]) and rows_data[r][c]:
                    headers[c] = rows_data[r][c]
                    break

    merged_cells_info = [
        {'row': k[0], 'col': k[1], 'rowspan': v.rowspan, 'colspan': v.colspan}
        for k, v in merged_map.items() if v.is_master
    ]

    records = []
    for row_idx, row_data in enumerate(rows_data[1:], start=2):
        record: dict = {'_row': row_idx, '_cells': {}}
        for col_idx, value in enumerate(row_data, start=1):
            key = (row_idx, col_idx)
            cell_info = merged_map.get(key)
            header_name = headers[col_idx - 1] if col_idx <= len(headers) else f'col_{col_idx}'

            if cell_info and not cell_info.is_master:
                mr, mc = cell_info.min_row, cell_info.min_col
                value = rows_data[mr - 1][mc - 1] if mr <= len(rows_data) and mc <= len(rows_data[mr - 1]) else ''

            cell_data: dict = {'value': value, 'col': col_idx, 'header': header_name}
            if cell_info and cell_info.is_master:
                cell_data.update(rowspan=cell_info.rowspan, colspan=cell_info.colspan, merged=True)
            elif cell_info and not cell_info.is_master:
                cell_data.update(merged=True, skipped=True)
            else:
                cell_data['merged'] = False
            record['_cells'][str(col_idx)] = cell_data
            record[f"{header_name}_col{col_idx}"] = value
        records.append(record)

    return json.dumps(
        {
            'metadata': {
                'sheet': sheet_name, 'total_rows': len(records),
                'total_columns': len(headers), 'headers': headers,
                'merged_cells': merged_cells_info,
            },
            'data': records,
        },
        ensure_ascii=False, indent=2,
    )
