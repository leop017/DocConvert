import json
import os
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from docconvert.converters import ExcelConverter
from docconvert.models import MergeInfo


def _make_xlsx(path, sheets):
    """Create an xlsx file. sheets is dict of {name: list_of_lists}."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(path)
    return path


def _make_xlsx_with_merges(path, sheet_name, rows, merges):
    """Create an xlsx file with merged cells. merges is list of 'A1:B1' strings."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title=sheet_name)
    for row in rows:
        ws.append(row)
    for m in merges:
        ws.merge_cells(m)
    wb.save(path)
    return path


class TestExcelConverterHTML(unittest.TestCase):
    """0/False must not render as &nbsp; in HTML header/body."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "t.xlsx")
        _make_xlsx(self.xlsx, {
            "Sheet1": [
                ["A", "B", "C"],
                [0, False, "x"],
                [1, True, "y"],
            ],
        })
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_html_does_not_render_zero_as_nbsp(self):
        c = ExcelConverter()
        results, errors = c.convert(self.xlsx, "html", self.out_dir)
        self.assertEqual(errors, [])
        self.assertEqual(len(results), 1)
        content = Path(results[0][1]).read_text(encoding="utf-8")
        # Each of the four 0/False cells should appear as text, not &nbsp;
        self.assertIn(">0<", content)
        self.assertIn(">False<", content)
        self.assertIn(">1<", content)
        self.assertIn(">True<", content)
        # Sanity: &nbsp; should not appear in the data rows (header is non-empty)
        body = content.split("<tbody>")[1].split("</tbody>")[0]
        self.assertNotIn("&nbsp;", body)

    def test_html_header_preserves_zero_and_false(self):
        c = ExcelConverter()
        results, _ = c.convert(self.xlsx, "html", self.out_dir)
        content = Path(results[0][1]).read_text(encoding="utf-8")
        # Headers are "A", "B", "C" — all non-empty so &nbsp; shouldn't appear
        self.assertNotIn("&nbsp;", content)


class TestExcelConverterMarkdownEnhanced(unittest.TestCase):
    """Enhanced MD path must preserve 0/False."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "t.xlsx")
        _make_xlsx(self.xlsx, {
            "Sheet1": [
                ["A", "B"],
                [0, False],
            ],
        })
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_enhanced_md_preserves_zero_and_false(self):
        c = ExcelConverter()
        results, errors = c.convert(
            self.xlsx, "md", self.out_dir, enhanced_md=True
        )
        self.assertEqual(errors, [])
        content = Path(results[0][1]).read_text(encoding="utf-8")
        self.assertIn("0", content)
        self.assertIn("False", content)

    def test_standard_md_runs_without_error(self):
        c = ExcelConverter()
        results, errors = c.convert(
            self.xlsx, "md", self.out_dir, enhanced_md=False
        )
        self.assertEqual(errors, [])
        self.assertTrue(Path(results[0][1]).exists())


class TestExcelConverterMergedCells(unittest.TestCase):
    """Merged cells should render with rowspan/colspan."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "merged.xlsx")
        _make_xlsx_with_merges(
            self.xlsx, "Sheet1",
            rows=[
                ["Header1", "Header1", "Header2"],
                ["A", "B", "C"],
                ["D", "E", "F"],
            ],
            merges=["A1:B1"],
        )
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_merged_master_has_rowspan_or_colspan(self):
        c = ExcelConverter()
        results, _ = c.convert(self.xlsx, "html", self.out_dir)
        content = Path(results[0][1]).read_text(encoding="utf-8")
        self.assertIn("colspan=\"2\"", content)

    def test_merged_map_build(self):
        c = ExcelConverter()
        ranges = c._load_merged_cache(self.xlsx, ".xlsx", ["Sheet1"])
        self.assertIsNotNone(ranges)
        self.assertIn("Sheet1", ranges)
        self.assertEqual(len(ranges["Sheet1"]), 1)
        m = c._build_merged_map(ranges["Sheet1"])
        master = m.get((1, 1))
        self.assertIsNotNone(master)
        self.assertTrue(master.is_master)
        self.assertEqual(master.colspan, 2)
        slave = m.get((1, 2))
        self.assertIsNotNone(slave)
        self.assertFalse(slave.is_master)

    def test_missing_sheet_yields_empty_list(self):
        """A missing sheet name must yield an empty list, not abort the whole cache."""
        c = ExcelConverter()
        cache = c._load_merged_cache(
            self.xlsx, ".xlsx", ["Sheet1", "NoSuchSheet"]
        )
        self.assertIsNotNone(cache)
        self.assertEqual(len(cache["NoSuchSheet"]), 0)
        self.assertEqual(len(cache["Sheet1"]), 1)


class TestExcelConverterXlsCache(unittest.TestCase):
    """``_load_merged_cache_xls`` must mirror the xlsx path's per-sheet
    defensive handling: a missing sheet name returns an empty list rather
    than aborting the whole cache.
    """

    def _make_xls(self, path, sheets):
        import xlwt
        wb = xlwt.Workbook()
        for name, rows in sheets.items():
            ws = wb.add_sheet(name)
            for r_idx, row in enumerate(rows):
                for c_idx, value in enumerate(row):
                    ws.write(r_idx, c_idx, value)
        wb.save(path)

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xls = os.path.join(self.tmp.name, "t.xls")
        self._make_xls(self.xls, {
            "Alpha": [["a"], [1]],
            "Beta":  [["b"], [2]],
        })

    def test_missing_sheet_yields_empty_list(self):
        """xls path must also tolerate missing sheet names."""
        c = ExcelConverter()
        cache = c._load_merged_cache_xls(
            self.xls, ["Alpha", "NoSuchSheet"]
        )
        self.assertIsNotNone(cache)
        self.assertEqual(len(cache["NoSuchSheet"]), 0)
        self.assertIn("Alpha", cache)

    def test_all_real_sheets(self):
        c = ExcelConverter()
        cache = c._load_merged_cache_xls(self.xls, ["Alpha", "Beta"])
        self.assertIn("Alpha", cache)
        self.assertIn("Beta", cache)


class TestExcelConverterJSON(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "t.xlsx")
        _make_xlsx(self.xlsx, {
            "Sheet1": [
                ["A", "B"],
                [0, 1],
            ],
        })
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_json_has_metadata_and_data(self):
        c = ExcelConverter()
        results, errors = c.convert(self.xlsx, "json", self.out_dir)
        self.assertEqual(errors, [])
        content = Path(results[0][1]).read_text(encoding="utf-8")
        parsed = json.loads(content)
        self.assertIn("metadata", parsed)
        self.assertIn("data", parsed)
        self.assertEqual(parsed["metadata"]["sheet"], "Sheet1")
        self.assertEqual(parsed["metadata"]["total_rows"], 1)
        self.assertEqual(parsed["metadata"]["headers"], ["A", "B"])

    def test_json_preserves_zero_value(self):
        c = ExcelConverter()
        results, _ = c.convert(self.xlsx, "json", self.out_dir)
        parsed = json.loads(Path(results[0][1]).read_text(encoding="utf-8"))
        first_row = parsed["data"][0]
        first_cell = first_row["_cells"]["1"]
        # Values pass through safe_str which stringifies numbers
        self.assertEqual(first_cell["value"], "0")
        self.assertEqual(first_row["A_col1"], "0")


class TestExcelConverterSheetSelection(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "multi.xlsx")
        _make_xlsx(self.xlsx, {
            "First":  [["f"], [1]],
            "Second": [["s"], [2]],
            "Third":  [["t"], [3]],
        })
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_sheets_param_limits_output(self):
        c = ExcelConverter()
        results, _ = c.convert(
            self.xlsx, "html", self.out_dir, sheets=["First", "Third"]
        )
        names = sorted(r[0] for r in results)
        self.assertEqual(len(results), 2)
        self.assertIn("multi_First.html", names)
        self.assertIn("multi_Third.html", names)
        self.assertNotIn("multi_Second.html", names)

    def test_nonexistent_sheet_reported_as_error(self):
        c = ExcelConverter()
        results, errors = c.convert(
            self.xlsx, "html", self.out_dir, sheets=["Nope"]
        )
        self.assertEqual(results, [])
        self.assertEqual(len(errors), 1)
        self.assertEqual(errors[0][0], "Nope")

    def test_sheets_param_does_not_load_unrequested_sheets(self):
        # When sheets= is given, pd.read_excel must receive the explicit
        # list (not None), so the converter never touches unrequested
        # sheets. Verified by patching pd.read_excel to capture the
        # actual sheet_name argument.
        from unittest.mock import patch
        c = ExcelConverter()
        with patch("docconvert.converters.excel.pd.read_excel") as mock_read:
            mock_read.return_value = {"First": None}
            c.convert(self.xlsx, "html", self.out_dir, sheets=["First"])
        # First positional/keyword call must be sheet_name=["First"]
        # (not None, which would have loaded every sheet)
        call_kwargs = mock_read.call_args.kwargs
        self.assertEqual(call_kwargs.get("sheet_name"), ["First"])
        self.assertNotEqual(call_kwargs.get("sheet_name"), None)

    def test_merged_cache_only_built_for_requested_sheets(self):
        # Regression: when sheets= is given, _load_merged_cache must
        # receive only the requested sheet names — not the full list.
        # This avoids opening the workbook an extra time for sheets
        # that will never be converted.
        from unittest.mock import patch
        c = ExcelConverter()
        with patch.object(c, "_load_merged_cache") as mock_cache:
            mock_cache.return_value = None
            with patch("docconvert.converters.excel.pd.read_excel") as mock_read:
                mock_read.return_value = {"First": None}
                c.convert(self.xlsx, "html", self.out_dir, sheets=["First"])
        # _load_merged_cache was called once, with sheet_names=["First"]
        mock_cache.assert_called_once()
        args, kwargs = mock_cache.call_args
        self.assertEqual(args[2], ["First"])


class TestExcelConverterCancellation(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "multi.xlsx")
        _make_xlsx(self.xlsx, {
            f"S{i}": [[i], [i]] for i in range(5)
        })
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_cancel_check_stops_processing(self):
        c = ExcelConverter()

        def cancel_after_first():
            c._cancel_called = getattr(c, "_cancel_called", 0) + 1
            return c._cancel_called > 1

        c.cancel_check = cancel_after_first
        results, _ = c.convert(self.xlsx, "html", self.out_dir)
        # Should convert at most 2 sheets before bailing out
        self.assertLess(len(results), 5)


class TestExcelConverterErrorPaths(unittest.TestCase):
    """load_workbook failure must not raise NameError."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_corrupted_xlsx_returns_none_no_nameerror(self):
        fake = os.path.join(self.tmp.name, "fake.xlsx")
        with open(fake, "wb") as f:
            f.write(b"not a real xlsx file")
        c = ExcelConverter()
        result = c._load_merged_cache(fake, ".xlsx", ["AnySheet"])
        self.assertIsNone(result)

    def test_unsupported_extension_returns_none(self):
        fake = os.path.join(self.tmp.name, "data.txt")
        Path(fake).write_text("hello")
        c = ExcelConverter()
        result = c._load_merged_cache(fake, ".txt", ["AnySheet"])
        self.assertIsNone(result)

    def test_invalid_output_format_reported_as_error(self):
        xlsx = os.path.join(self.tmp.name, "t.xlsx")
        _make_xlsx(xlsx, {"S": [["A"], ["x"]]})
        c = ExcelConverter()
        results, errors = c.convert(xlsx, "xml", self.out_dir)
        self.assertEqual(results, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("不支持的输出格式", errors[0][1])

    def test_empty_sheet_reported_as_error(self):
        xlsx = os.path.join(self.tmp.name, "empty.xlsx")
        wb = Workbook()
        wb.active.append([])
        wb.save(xlsx)
        c = ExcelConverter()
        results, errors = c.convert(xlsx, "html", self.out_dir)
        self.assertEqual(results, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("工作表为空", errors[0][1])


class TestExcelConverterAllSheets(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "t.xlsx")
        _make_xlsx(self.xlsx, {
            "A": [["a"], [1]],
            "B": [["b"], [2]],
        })
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_default_converts_all_sheets(self):
        c = ExcelConverter()
        results, errors = c.convert(self.xlsx, "html", self.out_dir)
        self.assertEqual(errors, [])
        self.assertEqual(len(results), 2)


class TestExcelConverterMergeFilter(unittest.TestCase):
    """Merge ranges that span rows dropped by dropna(how='all') must be
    filtered out, otherwise the rowspan would extend into a row that is
    not part of the merge in the source workbook.
    """

    def _make_xlsx_with_gap(self, path):
        """Header (row 1), data row 2, EMPTY row 3, data row 4.
        Merge A1:B1 spans header columns 1-2 (df index N/A, header) — KEPT.
        Merge A2:A4 spans column A across rows 2, 3, 4 (df index 1 dropped) — DROPPED.

        Layout:
          | H1        | H2 |
          | Section   | 2  |   <- row 2: A2 is merge slave, B2 has data
          | (empty)   |    |   <- row 3: gap
          | Section   | 4  |   <- row 4: A4 is merge slave, B4 has data
        """
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="S")
        ws.cell(row=1, column=1, value="H1")
        ws.cell(row=1, column=2, value="H2")
        ws.cell(row=2, column=2, value=2)  # B2 — keeps value
        # row 3 left empty (the gap)
        ws.cell(row=4, column=2, value=4)  # B4 — keeps value
        ws.merge_cells("A1:B1")   # header, 2 columns
        ws.merge_cells("A2:A4")   # data, 3 rows, crosses gap
        wb.save(path)

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "gap.xlsx")
        self._make_xlsx_with_gap(self.xlsx)
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_merges_spanning_gap_dropped(self):
        c = ExcelConverter()
        results, _ = c.convert(self.xlsx, "html", self.out_dir)
        content = Path(results[0][1]).read_text(encoding="utf-8")
        # A1:B1 (header merge, 2 columns) should be kept and render with colspan=2
        self.assertIn('colspan="2"', content,
                      "Header merge should still render with colspan=2")
        # A2:A4 (3-row merge crossing the gap) should be dropped
        self.assertNotIn('rowspan="3"', content,
                         "Merge that spans a dropped row should be removed")

    def test_filter_helper_keeps_header_merges_with_intact_legs(self):
        """Header-anchored merges (master in workbook row 1) are kept
        when none of their legs landed on a dropped row.
        """
        from docconvert.converters.excel import _XlsMergeRange
        m1 = _XlsMergeRange(1, 1, 1, 2)  # A1:A2  (header vertical)
        m2 = _XlsMergeRange(2, 1, 2, 4)  # B1:B4  (header vertical)
        # No rows dropped -> both survive.
        kept = ExcelConverter._filter_merges_for_dropped_rows(
            [m1, m2], dropped_df_indices=set()
        )
        self.assertEqual(len(kept), 2)

    def test_filter_helper_drops_header_merge_spanning_dropped_leg(self):
        """A header-anchored vertical merge whose leg lands on a dropped
        row MUST be dropped. Keeping it would let the master's rowspan
        swallow the independent cell that shifted up into the gap
        (silent data loss).
        """
        from docconvert.converters.excel import _XlsMergeRange
        m_keep = _XlsMergeRange(1, 1, 1, 2)  # A1:A2 legs df -1,0 -> intact
        m_drop = _XlsMergeRange(2, 1, 2, 4)  # B1:B4 legs df -1,0,1,2
        # dropna drops df index 2 (workbook row 4), a leg of m_drop.
        kept = ExcelConverter._filter_merges_for_dropped_rows(
            [m_keep, m_drop], dropped_df_indices={2}
        )
        self.assertEqual(len(kept), 1)
        self.assertEqual(kept[0].bounds, (1, 1, 1, 2))

    def test_filter_helper_drops_body_merges_spanning_drops(self):
        """Merges whose master is in the data area (workbook row >= 2)
        and whose legs include a dropped row must be removed, otherwise
        the renderer would emit a rowspan that jumps over the gap.
        """
        c = ExcelConverter()
        from docconvert.converters.excel import _XlsMergeRange
        m_keep = _XlsMergeRange(1, 2, 1, 2)  # A2:A2 (single row, no span)
        m_drop = _XlsMergeRange(1, 2, 1, 4)  # A2:A4 (spans dropped row 3)
        kept = ExcelConverter._filter_merges_for_dropped_rows(
            [m_keep, m_drop], dropped_df_indices={1}  # workbook row 3
        )
        self.assertEqual(len(kept), 1)
        self.assertEqual(kept[0].bounds, (1, 2, 1, 2))

    def test_filter_helper_keeps_all_when_no_drops(self):
        c = ExcelConverter()
        from docconvert.converters.excel import _XlsMergeRange
        m1 = _XlsMergeRange(1, 1, 1, 2)
        m2 = _XlsMergeRange(2, 1, 2, 4)
        kept = ExcelConverter._filter_merges_for_dropped_rows(
            [m1, m2], dropped_df_indices=set()
        )
        self.assertEqual(len(kept), 2)

    def test_filter_helper_keeps_intra_drop_merge(self):
        # Merge that does NOT touch a dropped row is preserved.
        from docconvert.converters.excel import _XlsMergeRange
        m1 = _XlsMergeRange(1, 1, 1, 2)
        kept = ExcelConverter._filter_merges_for_dropped_rows(
            [m1], dropped_df_indices={5}
        )
        self.assertEqual(len(kept), 1)


class TestExcelConverterHeaderMergeInteriorDrop(unittest.TestCase):
    """Regression: a header-anchored vertical merge (e.g. A1:A3) whose
    interior row is fully empty gets removed by dropna. The surviving
    independent cell below (A4) shifts up into the merge's rendered span
    and was silently masked as a merge slave -> data loss.
    """

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "hdr.xlsx")
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("S")
        ws.cell(row=1, column=1, value="HEAD")
        ws.cell(row=1, column=2, value="H2")
        ws.cell(row=2, column=2, value="r2")   # row 2: B has data, A2 slave
        # row 3 fully empty -> dropped by dropna
        ws.cell(row=4, column=1, value="SEPARATE")  # independent cell
        ws.cell(row=4, column=2, value="r4")
        ws.merge_cells("A1:A3")
        wb.save(self.xlsx)
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_independent_cell_below_gap_not_lost(self):
        c = ExcelConverter()
        res, _ = c.convert(self.xlsx, "html", self.out_dir)
        content = Path(res[0][1]).read_text(encoding="utf-8")
        tbody = content.split("<tbody>")[1].split("</tbody>")[0]
        self.assertIn("SEPARATE", tbody,
                      "Independent A4 cell must survive the interior-row drop")

    def test_json_output_keeps_independent_cell(self):
        c = ExcelConverter()
        res, _ = c.convert(self.xlsx, "json", self.out_dir)
        content = Path(res[0][1]).read_text(encoding="utf-8")
        self.assertIn("SEPARATE", content)


class TestExcelConverterMergeAfterBlankRow(unittest.TestCase):
    """Regression: a merge whose rows sit *after* a fully-empty spacer
    row must render at its correct post-dropna position. Before the fix
    the workbook row numbers were used verbatim while the table was
    re-indexed, so the merge drifted onto the wrong row (and even
    swallowed the wrong cell's value).
    """

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "gap.xlsx")
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("S")
        ws.cell(1, 1, "H1")
        ws.cell(1, 2, "H2")
        ws.cell(2, 1, "row2A")
        ws.cell(2, 2, "row2B")
        # row 3 fully empty -> dropped by dropna
        ws.cell(4, 1, "MERGED")
        ws.cell(4, 2, "X")  # A4:B4 after the gap
        ws.cell(5, 1, "row5A")
        ws.cell(5, 2, "row5B")
        ws.merge_cells("A4:B4")
        wb.save(self.xlsx)
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_html_merge_lands_on_correct_row(self):
        c = ExcelConverter()
        res, err = c.convert(self.xlsx, "html", self.out_dir)
        self.assertEqual(err, [])
        html = Path(res[0][1]).read_text(encoding="utf-8")
        tbody = html.split("<tbody>")[1].split("</tbody>")[0]
        # The merge belongs on rendered row 3, carrying the "MERGED" value.
        self.assertIn(
            'data-row="3"><td data-row="3" data-col="1" colspan="2"', tbody,
        )
        self.assertIn("MERGED", tbody)
        # The merged cell must NOT appear on the last row (row 5).
        self.assertNotIn('data-row="5" colspan="2"', tbody)
        # row5A stays in its own plain cell.
        self.assertIn("<td>row5A</td>", tbody)

    def test_json_merge_uses_remapped_row(self):
        c = ExcelConverter()
        res, err = c.convert(self.xlsx, "json", self.out_dir)
        self.assertEqual(err, [])
        data = json.loads(Path(res[0][1]).read_text(encoding="utf-8"))
        merged = [m for m in data["metadata"]["merged_cells"]
                  if m["colspan"] == 2]
        self.assertEqual(len(merged), 1)
        self.assertEqual(merged[0]["row"], 3)  # rendered row, not workbook 4
        self.assertEqual(merged[0]["col"], 1)
        # The merged master cell carries the correct value.
        cell = data["data"][1]["_cells"]["1"]  # _row 3 -> data index 1
        self.assertEqual(cell["value"], "MERGED")


class TestExcelConverterMarkdownPipeEscape(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, 't.xlsx')
        _make_xlsx(self.xlsx, {
            'S': [
                ['h1'],
                ['| pipe | text'],
            ],
        })
        self.out_dir = Path(self.tmp.name) / 'out'
        self.out_dir.mkdir()

    def test_pipe_escaped_in_standard_md(self):
        c = ExcelConverter()
        res, _ = c.convert(self.xlsx, 'md', self.out_dir, enhanced_md=False)
        content = Path(res[0][1]).read_text(encoding='utf-8')
        # Each pipe inside a cell must be escaped so the rendered
        # table keeps its column count (1 col, not 5).
        self.assertIn('\\| pipe \\| text', content)
        # The data row must declare exactly 1 column.
        self.assertIn('| \\| pipe \\| text |', content)

    def test_pipe_escaped_in_enhanced_md(self):
        c = ExcelConverter()
        res, _ = c.convert(self.xlsx, 'md', self.out_dir, enhanced_md=True)
        content = Path(res[0][1]).read_text(encoding='utf-8')
        self.assertIn('\\| pipe \\| text', content)


class TestExcelConverterMarkdownNewlineEscape(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, 't.xlsx')
        # Use a real newline character so pd / openpyxl preserve it.
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title='S')
        ws.append(['h1'])
        ws.append(['line1' + chr(10) + 'line2'])
        wb.save(self.xlsx)
        self.out_dir = Path(self.tmp.name) / 'out'
        self.out_dir.mkdir()

    def test_standard_md_preserves_newline_as_br(self):
        # df.to_html escapes the newline as a literal backslash-n
        # sequence. _escape_table_cells must unescape it so the
        # final markdown contains <br> instead of literal text.
        c = ExcelConverter()
        res, _ = c.convert(self.xlsx, 'md', self.out_dir, enhanced_md=False)
        content = Path(res[0][1]).read_text(encoding='utf-8')
        self.assertIn('line1<br>line2', content)
        # Sanity: no literal backslash-n survives in the rendered cell.
        self.assertNotIn('line1\\\\nline2', content)

    def test_enhanced_md_escapes_pipe_with_newline(self):
        # Enhanced path uses rows_data (preserves newlines); the
        # escape helper applies <br> substitution.
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title='S')
        ws.append(['h1'])
        ws.append(['| a' + chr(10) + 'b |'])
        xlsx = os.path.join(self.tmp.name, 'p.xlsx')
        wb.save(xlsx)
        c = ExcelConverter()
        res, _ = c.convert(xlsx, 'md', self.out_dir, enhanced_md=True)
        content = Path(res[0][1]).read_text(encoding='utf-8')
        self.assertIn('\\| a', content)
        self.assertIn('b \\|', content)

    def test_enhanced_md_preserves_newline_as_br(self):
        # Regression: the enhanced path escaped the newline to a literal
        # ``<br>`` that the intermediate BeautifulSoup parse turned into a
        # real element, which markdownify then collapsed to a space. The
        # line break must survive as ``<br>`` text, like the standard path.
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title='S')
        ws.append(['h1'])
        ws.append(['line1' + chr(10) + 'line2'])
        xlsx = os.path.join(self.tmp.name, 'nl.xlsx')
        wb.save(xlsx)
        c = ExcelConverter()
        res, _ = c.convert(xlsx, 'md', self.out_dir, enhanced_md=True)
        content = Path(res[0][1]).read_text(encoding='utf-8')
        self.assertIn('line1<br>line2', content)
        # The two lines must not be merged into a single space.
        self.assertNotIn('line1 line2', content)


class TestExcelConverterMergeSkippedRowGuard(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, 't.xlsx')
        # Workbook rows:
        #   row 1: header
        #   row 2: 'data' | 2     (master of A2:A4)
        #   row 3: 'data' | 4     (slave of A2:A4, value cleared
        #                          by openpyxl merge, col 2 kept)
        #   row 4: None  | None   (slave of A2:A4, fully empty)
        # pd.read_excel silently drops the trailing fully-empty row
        # so the rendered df has only 2 data rows. The merge
        # guard must drop the merge to avoid a rowspan that no
        # tbody row covers.
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title='S')
        ws.append(['h1', 'h2'])
        ws.append(['data', 2])
        ws.append(['data', 4])
        ws.merge_cells('A2:A4')
        wb.save(self.xlsx)
        self.out_dir = Path(self.tmp.name) / 'out'
        self.out_dir.mkdir()

    def test_rowspan_into_skipped_row_is_dropped(self):
        c = ExcelConverter()
        res, _ = c.convert(self.xlsx, 'html', self.out_dir)
        content = Path(res[0][1]).read_text(encoding='utf-8')
        tbody = content.split('<tbody>')[1].split('</tbody>')[0]
        tbody_rows = tbody.count('<tr')
        self.assertEqual(tbody_rows, 2)
        # The guard drops the entire merge, so no rowspan=3 appears.
        self.assertNotIn('rowspan="3"', content)
        self.assertNotIn('rowspan=' + chr(34) + '3' + chr(34), content)

class TestExcelConverterHeaderVerticalMerge(unittest.TestCase):
    """Regression: a header-anchored vertical merge (e.g. A1:A3) cannot be
    expressed with a ``<th rowspan>`` that crosses the thead/tbody boundary
    — browsers clamp it to the thead, so the body rows lost their first
    column and the table misaligned. It must be flattened: a normal header
    cell plus placeholder cells in the covered body columns.
    """

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "hdr.xlsx")
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("S")
        ws.append(["HEAD", "B"])
        ws.append(["x", 2])
        ws.append(["", 3])
        ws.merge_cells("A1:A3")
        wb.save(self.xlsx)
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_body_rows_stay_aligned(self):
        c = ExcelConverter()
        res, errs = c.convert(self.xlsx, "html", self.out_dir)
        self.assertEqual(errs, [])
        content = Path(res[0][1]).read_text(encoding="utf-8")
        tbody = content.split("<tbody>")[1].split("</tbody>")[0]
        self.assertEqual(tbody.count("<tr"), 2)
        # Both columns must survive in every body row (placeholder for col 1)
        self.assertEqual(tbody.count("<td"), 4)

    def test_thead_has_no_rowspan_and_keeps_header(self):
        c = ExcelConverter()
        res, _ = c.convert(self.xlsx, "html", self.out_dir)
        content = Path(res[0][1]).read_text(encoding="utf-8")
        thead = content.split("<thead>")[1].split("</thead>")[0]
        self.assertIn("HEAD", thead)
        self.assertNotIn("rowspan", thead)


class TestExcelConverterHeaderMergeColspan(unittest.TestCase):
    """Regression: a header-anchored merge that spans BOTH rows and columns
    (e.g. A1:C2) must keep its colspan when flattened. Previously the
    flattened ``<th>`` dropped colspan, leaving the thead with fewer columns
    than the tbody and misaligning the remaining header cells.
    """

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "hdr_span.xlsx")
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("S")
        ws.append(["col1", "col2", "col3", "col4"])
        ws.merge_cells("A1:C2")
        ws["A1"] = "BIG HEADER"
        ws.append(["a", "b", "c", "d"])
        ws.append(["e", "f", "g", "h"])
        ws.append(["i", "j", "k", "l"])
        wb.save(self.xlsx)
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_header_merge_keeps_colspan(self):
        c = ExcelConverter()
        results, errors = c.convert(self.xlsx, "html", self.out_dir)
        self.assertEqual(errors, [])
        content = Path(results[0][1]).read_text(encoding="utf-8")
        thead = content.split("<thead>")[1].split("</thead>")[0]
        # Flattened header still spans the merge's columns.
        self.assertIn('colspan="3"', thead)
        # The trailing header cell stays on its own column.
        self.assertIn('data-col="4"', thead)

    def test_body_rows_stay_aligned(self):
        c = ExcelConverter()
        results, _ = c.convert(self.xlsx, "html", self.out_dir)
        content = Path(results[0][1]).read_text(encoding="utf-8")
        tbody = content.split("<tbody>")[1].split("</tbody>")[0]
        # Every body row keeps all 4 columns.
        self.assertEqual(tbody.count("<td"), 4 * 3)


class TestExcelConverterSheetNameCollision(unittest.TestCase):
    """Regression: sheet names that clean to the same filename (e.g.
    ``Q"1`` and ``Q<1`` → ``Q_1``) must not share an output path, or the
    later sheet would silently overwrite the earlier one.
    """

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "book.xlsx")
        wb = Workbook()
        wb.remove(wb.active)
        ws1 = wb.create_sheet('Q"1')
        ws1.append(["h"])
        ws1.append(["DATA1"])
        ws2 = wb.create_sheet("Q<1")
        ws2.append(["h"])
        ws2.append(["DATA2"])
        wb.save(self.xlsx)
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_colliding_sheet_names_get_unique_outputs(self):
        c = ExcelConverter()
        results, errors = c.convert(self.xlsx, "html", self.out_dir)
        self.assertEqual(errors, [])
        self.assertEqual(len(results), 2)
        paths = {p for _, p in results}
        self.assertEqual(len(paths), 2, "Colliding sheet names must not overwrite")
        blob = "".join(Path(p).read_text(encoding="utf-8") for p in paths)
        self.assertIn("DATA1", blob)
        self.assertIn("DATA2", blob)

    def test_overwrite_check_matches_unique_outputs(self):
        from docconvert.controller import ConversionController
        c = ExcelConverter()
        c.convert(self.xlsx, "html", self.out_dir)
        controller = ConversionController()
        existing = controller.check_overwrite_paths(
            [self.xlsx], "html", output_dir=str(self.out_dir)
        )
        self.assertEqual(len(set(existing)), 2)


class TestExcelConverterEmptyHeaderFallback(unittest.TestCase):
    """When all header values in rows_data are falsy, _generate_json_data
    must fall back to the first non-empty value found in each column.
    """

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "t.xlsx")
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("S")
        ws.append(["", ""])   # all-empty header row
        ws.append(["val_a", "val_b"])
        ws.append(["x", "y"])
        wb.save(self.xlsx)
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_empty_header_falls_back_to_first_row(self):
        # pandas auto-names empty headers as "Unnamed: N", so we test the
        # fallback by calling _generate_json_data directly with falsy headers.
        from docconvert.converters.excel import _generate_json_data
        rows_data = [["", ""], ["val_a", "val_b"], ["x", "y"]]
        result = json.loads(_generate_json_data(rows_data, {}, "S"))
        self.assertEqual(result["metadata"]["headers"], ["val_a", "val_b"])


class TestExcelConverterSingleSheetNoDict(unittest.TestCase):
    """pd.read_excel with a single sheet and no sheets= param returns a
    DataFrame (not a dict). The converter must wrap it so downstream code
    sees a dict keyed by sheet name.
    """

    def test_single_sheet_returns_one_result(self):
        c = ExcelConverter()
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        xlsx = os.path.join(tmp.name, "t.xlsx")
        _make_xlsx(xlsx, {"Only": [["A"], ["1"]]})
        out = Path(tmp.name) / "out"
        out.mkdir()
        results, errors = c.convert(xlsx, "html", out)
        self.assertEqual(errors, [])
        self.assertEqual(len(results), 1)


class TestExcelConverterXlsImportError(unittest.TestCase):
    """Converting .xls without xlrd installed must raise RuntimeError."""

    def test_xls_without_xlrd_raises(self):
        import builtins
        from unittest.mock import patch
        c = ExcelConverter()
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        xls = os.path.join(tmp.name, "t.xls")
        with open(xls, "wb") as f:
            f.write(b"fake xls")
        out = Path(tmp.name) / "out"
        out.mkdir()
        # Mock builtins.__import__ to simulate xlrd not being installed
        real_import = builtins.__import__
        def mock_import(name, *args, **kwargs):
            if name == "xlrd":
                raise ImportError("No module named 'xlrd'")
            return real_import(name, *args, **kwargs)
        with patch.object(builtins, '__import__', mock_import):
            with self.assertRaises(RuntimeError) as ctx:
                c.convert(xls, "html", out)
            self.assertIn("xlrd", str(ctx.exception))


class TestExcelConverterMergeIteration(unittest.TestCase):
    """Directly test _iter_merge_legs and _cell_attrs for rowspan/colspan."""

    def test_iter_merge_legs_yields_all_cells(self):
        c = ExcelConverter()
        info = MergeInfo(
            rowspan=2, colspan=3, is_master=True,
            is_merged=True, min_row=1, min_col=1, max_row=2, max_col=3,
        )
        legs = list(c._iter_merge_legs(info))
        self.assertEqual(len(legs), 6)
        self.assertIn((1, 1), legs)
        self.assertIn((2, 3), legs)

    def test_cell_attrs_includes_rowspan(self):
        c = ExcelConverter()
        info = MergeInfo(
            rowspan=3, colspan=1, is_master=True,
            is_merged=True, min_row=1, min_col=1, max_row=3, max_col=1,
        )
        attrs = c._cell_attrs(1, 1, info)
        self.assertIn('rowspan="3"', attrs)
        self.assertIn('data-rowspan="3"', attrs)

    def test_cell_attrs_includes_colspan(self):
        c = ExcelConverter()
        info = MergeInfo(
            rowspan=1, colspan=2, is_master=True,
            is_merged=True, min_row=1, min_col=1, max_row=1, max_col=2,
        )
        attrs = c._cell_attrs(1, 1, info)
        self.assertIn('colspan="2"', attrs)
        self.assertIn('data-colspan="2"', attrs)


class TestExcelConverterEmptyRowsGuard(unittest.TestCase):
    """When a merge's max_row exceeds the post-dropna row count, the merge
    must be removed entirely (not left dangling).
    """

    def test_merge_beyond_dropna_rows_is_dropped(self):
        from docconvert.converters.excel import _XlsMergeRange
        c = ExcelConverter()
        merged_ranges = [
            _XlsMergeRange(min_col=1, min_row=1, max_col=1, max_row=3)
        ]
        # Workbook row 2 maps to df index 0; drop that row -> merge spans a dropped row
        dropped = {0}
        remapped = c._filter_merges_for_dropped_rows(merged_ranges, dropped)
        # Merge covering a dropped row must be removed
        self.assertEqual(len(remapped), 0)


class TestExcelConverterEnhancedMdMergeRendering(unittest.TestCase):
    """Enhanced MD path must correctly render merge cells (master + slave)."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = os.path.join(self.tmp.name, "t.xlsx")
        _make_xlsx_with_merges(
            self.xlsx, "S",
            rows=[
                ["M1", "M1", "B"],
                ["A", "B", "C"],
            ],
            merges=["A1:B1"],
        )
        self.out_dir = Path(self.tmp.name) / "out"
        self.out_dir.mkdir()

    def test_enhanced_md_rendered(self):
        c = ExcelConverter()
        results, errors = c.convert(self.xlsx, "md", self.out_dir, enhanced_md=True)
        self.assertEqual(errors, [])
        content = Path(results[0][1]).read_text(encoding="utf-8")
        self.assertIn("M1", content)
        self.assertIn("A", content)


class TestExcelConverterEscapeTableCells(unittest.TestCase):
    """Test _escape_table_cells covers the isinstance(Tag) guard path."""

    def test_pipe_escaped_via_escape_table_cells(self):
        c = ExcelConverter()
        from bs4 import BeautifulSoup
        html = '<table><tr><td>| pipe |</td></tr></table>'
        result = c._escape_table_cells(html)
        self.assertIn("\\| pipe \\|", result)


class TestExcelConverterColspanAttrs(unittest.TestCase):
    """HTML output with colspan merge must include both rowspan and colspan
    attrs when both are >1."""

    def test_colspan_only_attr(self):
        c = ExcelConverter()
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        xlsx = os.path.join(tmp.name, "t.xlsx")
        _make_xlsx_with_merges(
            xlsx, "S",
            rows=[["H1", "H1", "H2"], ["a", "b", "c"]],
            merges=["A1:B1"],
        )
        out = Path(tmp.name) / "out"
        out.mkdir()
        results, _ = c.convert(xlsx, "html", out)
        content = Path(results[0][1]).read_text(encoding="utf-8")
        self.assertIn('colspan="2"', content)
        self.assertIn('data-colspan="2"', content)
        # Must NOT have rowspan since it's only a colspan merge
        self.assertNotIn('rowspan="2"', content)


if __name__ == '__main__':
    unittest.main()
